from flask import (Flask, render_template, request,
                   redirect, session, flash, make_response, jsonify)
import smtplib
from email.message import EmailMessage
import joblib
import numpy as np
import secrets
from datetime import datetime, timedelta
from io import BytesIO
import sqlite3

# ── PDF export ──────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, PageBreak)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Excel export ─────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from db import (
    init_db, get_latest_prediction, get_patient_predictions,
    save_prediction, request_discharge, get_discharge_status,
    get_pending_discharges, update_discharge, get_all_predictions,
    get_patient_email, save_symptom, get_all_symptoms,
    get_flagged_symptom_count, get_all_discharge_history,
    acknowledge_care_plan, schedule_appointment,
    get_patient_appointments, get_all_appointments,
    prescribe_medication, get_patient_medications, get_all_medications,
    add_notification, get_notifications, mark_notifications_read,
    get_unread_count, get_analytics,
    save_vital_signs, get_patient_vitals, get_all_vitals,
    send_message, get_messages_for_patient,
    get_unread_message_count, mark_messages_read,
    trigger_sos, get_all_sos_alerts, resolve_sos, get_unresolved_sos_count,
    save_feedback, get_all_feedback, get_patient_report,
    log_action, get_audit_log, get_all_patients,
    get_user_by_email, update_user_password,
    get_user_profile, update_user_profile
)
from auth import register_user, login_user
from utils import get_care_plan

app = Flask(__name__)
app.secret_key = "secure_key"
init_db()

model = joblib.load("model/readmission_model.pkl")

SMTP_SERVER     = "smtp.gmail.com"
SMTP_PORT       = 587
SENDER_EMAIL    = "postdischargecareteam@gmail.com"
SENDER_PASSWORD = "xeda vipz beeq lgha"
reset_tokens    = {}

GEMINI_API_KEY  = "AIzaSyDF2YYMVt8qJ_Cu4g-yT7E7FsoiUG7dt-s"

def init_chatbot_db():
    conn = sqlite3.connect("database.db")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS chat_history (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id TEXT NOT NULL,
            message    TEXT NOT NULL,
            response   TEXT NOT NULL,
            timestamp  TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

init_chatbot_db()


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def send_email(to_email, subject, body):
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From']    = SENDER_EMAIL
        msg['To']      = to_email
        msg.set_content(body)
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.ehlo(); smtp.starttls()
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp.send_message(msg)
        return True
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False

def send_welcome_email(to, name, uid):
    return send_email(to, "Registration Successful",
        f"Hello {name},\n\nYour ID: {uid}\n\nRegards,\nPost Discharge Care Team")

def send_reset_email(to, name, link):
    return send_email(to, "Password Reset",
        f"Hello {name},\n\nReset link (30 mins):\n{link}\n\nRegards,\nPost Discharge Care Team")

def send_risk_alert_email(to, name, risk, prob, plan):
    return send_email(to, f"⚠️ Risk Alert: {risk}",
        f"Hello {name},\n\nRisk: {risk}\nProbability: {round(prob*100,2)}%\n\nCare Plan:\n{plan}\n\nRegards,\nPost Discharge Care Team")

def send_discharge_decision_email(to, name, status, remarks):
    e = "✅" if status == "APPROVED" else "❌"
    return send_email(to, f"{e} Discharge {status.capitalize()}",
        f"Hello {name},\n\nDecision: {status}\nRemarks: {remarks or 'None'}\n\nRegards,\nPost Discharge Care Team")

def send_symptom_alert_email(patient_name, fever, pain, breathing, notes):
    return send_email(SENDER_EMAIL, f"🚨 Symptom Alert — {patient_name}",
        f"Patient: {patient_name}\nFever: {fever}\nPain: {pain}/10\nBreathing: {breathing}\nNotes: {notes or 'None'}")

def send_sos_email(patient_name, patient_id, last_symptoms):
    return send_email(SENDER_EMAIL, f"🆘 EMERGENCY SOS — {patient_name}",
        f"EMERGENCY ALERT!\n\nPatient: {patient_name}\nID: {patient_id}\n\nLast Symptoms:\n{last_symptoms}\n\nPlease contact patient immediately!")

def send_appointment_email(to, name, date, time, reason):
    return send_email(to, "📅 Appointment Scheduled",
        f"Hello {name},\n\nDate: {date}\nTime: {time}\nReason: {reason}\n\nRegards,\nPost Discharge Care Team")

def send_medication_email(to, name, med, dosage, freq, duration, instr):
    return send_email(to, "💊 New Medication Prescribed",
        f"Hello {name},\n\nMedicine: {med}\nDosage: {dosage}\nFrequency: {freq}\nDuration: {duration}\nInstructions: {instr or 'As directed'}\n\nRegards,\nPost Discharge Care Team")

def send_vital_alert_email(patient_name, bp, hr, temp, bs):
    return send_email(SENDER_EMAIL, f"⚠️ Abnormal Vitals — {patient_name}",
        f"Patient: {patient_name}\nBP: {bp}\nHeart Rate: {hr}\nTemperature: {temp}°C\nBlood Sugar: {bs} mg/dL\n\nImmediate review required.")


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYTICS HELPER
# ═══════════════════════════════════════════════════════════════════════════════

def _get_analytics_data():
    analytics   = get_analytics()
    predictions = get_all_predictions()
    vitals      = get_all_vitals()
    symptoms    = get_all_symptoms()
    feedbacks   = get_all_feedback()
    patients    = get_all_patients()
    discharges  = get_all_discharge_history()

    risk_counts = {"High": 0, "Medium": 0, "Low": 0}
    for p in predictions:
        risk_val = p[3] if len(p) > 3 and p[3] in ["High", "Medium", "Low"] else p[2]
        risk_counts[risk_val] = risk_counts.get(risk_val, 0) + 1

    recent_preds = predictions[-20:]
    prob_labels  = []
    prob_values  = []

    for p in recent_preds:
        try:
            prob = float(p[3])
        except (ValueError, TypeError):
            try:
                prob = float(p[4])
            except (ValueError, TypeError, IndexError):
                prob = 0.0
        prob_values.append(round(prob * 100, 2))
        if len(p) > 6 and "20" in str(p[6]):
            prob_labels.append(str(p[6])[:10])
        else:
            prob_labels.append(str(p[5])[:10] if len(p) > 5 else "N/A")

    hr_vals, temp_vals, bs_vals = [], [], []
    for v in vitals:
        try: hr_vals.append(float(v[2]))
        except (ValueError, TypeError, IndexError): pass
        try: temp_vals.append(float(v[3]))
        except (ValueError, TypeError, IndexError): pass
        try: bs_vals.append(float(v[4]))
        except (ValueError, TypeError, IndexError): pass

    avg_hr   = round(sum(hr_vals)   / len(hr_vals),   1) if hr_vals   else 0
    avg_temp = round(sum(temp_vals) / len(temp_vals), 1) if temp_vals else 0
    avg_bs   = round(sum(bs_vals)   / len(bs_vals),   1) if bs_vals   else 0

    fever_yes = breathing_abn = high_pain = 0
    for s in symptoms:
        if "Yes" in str(s):
            fever_yes += 1
        b_val = str(s[4]) if len(s) > 4 else "Normal"
        if b_val in ["Yes", "No"] or b_val.isdigit():
            b_val = str(s[5]) if len(s) > 5 else "Normal"
        if b_val not in ["Normal", "None", ""]:
            breathing_abn += 1
        pain_val = 0
        try:
            pain_val = int(s[3])
        except (ValueError, TypeError, IndexError):
            try:
                pain_val = int(s[4])
            except (ValueError, TypeError, IndexError):
                pass
        if pain_val >= 7:
            high_pain += 1

    avg_rating   = (round(sum(f[2] for f in feedbacks) / len(feedbacks), 1)
                    if feedbacks else 0)
    dis_approved = sum(1 for d in discharges if d[3] == "APPROVED")
    dis_rejected = sum(1 for d in discharges if d[3] == "REJECTED")

    return {
        "total_patients":    len(patients),
        "total_predictions": len(predictions),
        "analytics":         analytics,
        "risk_counts":       risk_counts,
        "prob_labels":       prob_labels,
        "prob_values":       prob_values,
        "avg_hr":            avg_hr,
        "avg_temp":          avg_temp,
        "avg_bs":            avg_bs,
        "fever_yes":         fever_yes,
        "breathing_abn":     breathing_abn,
        "high_pain":         high_pain,
        "avg_rating":        avg_rating,
        "dis_approved":      dis_approved,
        "dis_rejected":      dis_rejected,
        "total_discharges":  len(discharges),
        "predictions":       predictions,
        "vitals":            vitals,
        "symptoms":          symptoms,
        "feedbacks":         feedbacks,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE ① — ANALYTICS DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/analytics-dashboard")
def analytics_dashboard():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')
    data = _get_analytics_data()
    log_action(session['user'], "VIEW_ANALYTICS_DASHBOARD")
    return render_template("analytics_dashboard.html", **data)


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE ② — EXPORT FULL REPORT AS PDF
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/export-report-pdf")
def export_report_pdf():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')

    data = _get_analytics_data()
    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=0.75*inch,  bottomMargin=0.75*inch)

    styles     = getSampleStyleSheet()
    BLUE       = colors.HexColor("#1a73e8")
    LBLUE      = colors.HexColor("#e8f0fe")
    GREEN      = colors.HexColor("#34a853")
    RED        = colors.HexColor("#ea4335")
    AMBER      = colors.HexColor("#fbbc04")
    DGRAY      = colors.HexColor("#3c4043")

    title_style = ParagraphStyle("Title2", parent=styles["Title"],
                                  textColor=BLUE, fontSize=20, spaceAfter=6)
    h1_style    = ParagraphStyle("H1", parent=styles["Heading1"],
                                  textColor=BLUE, fontSize=13, spaceAfter=4)
    h2_style    = ParagraphStyle("H2", parent=styles["Heading2"],
                                  textColor=DGRAY, fontSize=11, spaceAfter=3)
    body_style  = styles["Normal"]
    small_style = ParagraphStyle("Small", parent=styles["Normal"],
                                  fontSize=8, textColor=colors.grey)

    def section(title):
        return [Spacer(1, 14), Paragraph(title, h1_style),
                HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=6)]

    def kv_table(rows, col_widths=None):
        tbl = Table(rows, colWidths=col_widths or [2.5*inch, 3.5*inch])
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (0,-1), LBLUE),
            ('TEXTCOLOR',     (0,0), (0,-1), DGRAY),
            ('FONTNAME',      (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,0), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('RIGHTPADDING',  (0,0), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        return tbl

    def data_table(header, rows, col_widths=None):
        tbl = Table([header] + rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('RIGHTPADDING',  (0,0), (-1,-1), 6),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ]))
        return tbl

    story = []
    story.append(Spacer(1, 40))
    story.append(Paragraph("Post Discharge Care System", title_style))
    story.append(Paragraph("Comprehensive Analytics Report", h2_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"By: {session['name']} ({session['role'].title()})", small_style))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=20))

    story += section("Summary Overview")
    a = data["analytics"]
    kv_rows = [
        ["Total Patients Registered",   str(data["total_patients"])],
        ["Total Risk Predictions",       str(data["total_predictions"])],
        ["High Risk Patients",           str(data["risk_counts"].get("High",0))],
        ["Medium Risk Patients",         str(data["risk_counts"].get("Medium",0))],
        ["Low Risk Patients",            str(data["risk_counts"].get("Low",0))],
        ["Discharge Requests Approved",  str(data["dis_approved"])],
        ["Discharge Requests Rejected",  str(data["dis_rejected"])],
        ["Average Patient Rating",       f"{data['avg_rating']} / 5"],
        ["Average Heart Rate",           f"{data['avg_hr']} bpm"],
        ["Average Temperature",          f"{data['avg_temp']} °C"],
        ["Average Blood Sugar",          f"{data['avg_bs']} mg/dL"],
    ]
    story.append(kv_table(kv_rows))

    story += section("Risk Distribution")
    total_p   = data["total_predictions"] or 1
    risk_rows = [
        [lvl, str(data["risk_counts"].get(lvl,0)),
         f"{round(data['risk_counts'].get(lvl,0)/total_p*100,1)}%"]
        for lvl in ["High", "Medium", "Low"]
    ]
    story.append(data_table(["Risk Level","Count","Percentage"], risk_rows,
                             col_widths=[2*inch,1.5*inch,1.5*inch]))

    story += section("Symptom Highlights")
    story.append(kv_table([
        ["Patients with Fever",              str(data["fever_yes"])],
        ["Patients with Abnormal Breathing", str(data["breathing_abn"])],
        ["Patients with High Pain (>=7)",    str(data["high_pain"])],
    ]))

    story += section("Recent Risk Predictions (Last 20)")
    pred_rows = []
    for p in data["predictions"][-20:]:
        try:
            pred_rows.append([str(p[1]), str(p[2]),
                               f"{round(float(p[3])*100,1)}%", str(p[5])[:16]])
        except Exception:
            pred_rows.append([str(x)[:20] for x in p[:4]])
    if pred_rows:
        story.append(data_table(["Patient ID","Risk","Probability","Date"],
                                 pred_rows, col_widths=[2*inch,1.2*inch,1.5*inch,2*inch]))
    else:
        story.append(Paragraph("No prediction data available.", body_style))

    story += section("Recent Vitals (Last 15)")
    if data["vitals"]:
        vit_rows = [[str(v[1]),str(v[1]),str(v[2]),str(v[3]),str(v[4]),str(v[5])[:16]]
                    for v in data["vitals"][:15]]
        story.append(data_table(["Patient ID","BP","HR","Temp (°C)","B.Sugar","Date"],
                                 vit_rows,
                                 col_widths=[1.5*inch,1.2*inch,0.8*inch,1*inch,1*inch,1.8*inch]))
    else:
        story.append(Paragraph("No vitals data available.", body_style))

    story += section("Patient Feedback Summary")
    story.append(kv_table([
        ["Total Feedback Responses", str(len(data["feedbacks"]))],
        ["Average Rating",           f"{data['avg_rating']} / 5.0"],
    ]))
    if data["feedbacks"]:
        story.append(Spacer(1, 8))
        fb_data = [[str(f[1]), f"{f[2]}/5", str(f[3] or "—")[:60]]
                   for f in data["feedbacks"][:10]]
        story.append(data_table(["Patient ID","Rating","Comments"], fb_data,
                                 col_widths=[1.5*inch,1*inch,4.2*inch]))

    story.append(PageBreak())
    story.append(Spacer(1, 200))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        "Post Discharge Care Optimization System — Confidential Report",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    log_action(session['user'], "EXPORT_REPORT_PDF")
    response = make_response(buf.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = \
        f'attachment; filename=analytics_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE ③ — EXPORT FULL REPORT AS EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/export-report-excel")
def export_report_excel():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')

    data = _get_analytics_data()
    wb   = openpyxl.Workbook()

    BLUE_FILL  = PatternFill("solid", fgColor="1a73e8")
    LBLUE_FILL = PatternFill("solid", fgColor="e8f0fe")
    GREEN_FILL = PatternFill("solid", fgColor="34a853")
    RED_FILL   = PatternFill("solid", fgColor="ea4335")
    AMBER_FILL = PatternFill("solid", fgColor="fbbc04")
    GRAY_FILL  = PatternFill("solid", fgColor="f8f9fa")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="1a73e8", size=14)
    bold_font  = Font(bold=True, size=10)
    norm_font  = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color="DEE2E6"),
        right=Side(style='thin', color="DEE2E6"),
        top=Side(style='thin', color="DEE2E6"),
        bottom=Side(style='thin', color="DEE2E6"),
    )

    def set_header_row(ws, row_num, headers, widths=None):
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=row_num, column=col, value=h)
            cell.fill      = BLUE_FILL
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            if widths:
                ws.column_dimensions[get_column_letter(col)].width = widths[col-1]

    def style_data_row(ws, row_num, ncols, alt=False):
        for col in range(1, ncols + 1):
            cell           = ws.cell(row=row_num, column=col)
            cell.fill      = GRAY_FILL if alt else PatternFill()
            cell.font      = norm_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border

    def make_title(ws, text):
        ws["A1"] = text
        ws["A1"].font      = title_font
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  By: {session['name']}"
        ws["A2"].font = Font(italic=True, size=9, color="888888")
        ws.row_dimensions[1].height = 24

    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    make_title(ws1, "Post Discharge Care — Analytics Summary")
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20

    kpis = [
        ("Total Patients",             data["total_patients"],              BLUE_FILL),
        ("Total Risk Predictions",     data["total_predictions"],           BLUE_FILL),
        ("High Risk Patients",         data["risk_counts"].get("High",0),   RED_FILL),
        ("Medium Risk Patients",       data["risk_counts"].get("Medium",0), AMBER_FILL),
        ("Low Risk Patients",          data["risk_counts"].get("Low",0),    GREEN_FILL),
        ("Discharges Approved",        data["dis_approved"],                GREEN_FILL),
        ("Discharges Rejected",        data["dis_rejected"],                RED_FILL),
        ("Avg Patient Rating (/ 5)",   data["avg_rating"],                  LBLUE_FILL),
        ("Avg Heart Rate (bpm)",       data["avg_hr"],                      LBLUE_FILL),
        ("Avg Temperature (°C)",       data["avg_temp"],                    LBLUE_FILL),
        ("Avg Blood Sugar (mg/dL)",    data["avg_bs"],                      LBLUE_FILL),
        ("Fever Reported",             data["fever_yes"],                   AMBER_FILL),
        ("Abnormal Breathing",         data["breathing_abn"],               AMBER_FILL),
        ("High Pain Level (>=7)",      data["high_pain"],                   AMBER_FILL),
    ]
    start = 4
    ws1.cell(row=start, column=1).fill = BLUE_FILL
    ws1.cell(row=start, column=2).fill = BLUE_FILL
    ws1.cell(row=start, column=1).font = hdr_font
    ws1.cell(row=start, column=2).font = hdr_font
    ws1.cell(row=start, column=1, value="Metric")
    ws1.cell(row=start, column=2, value="Value")

    for i, (label, value, fill) in enumerate(kpis, start=1):
        r  = start + i
        lc = ws1.cell(row=r, column=1, value=label)
        vc = ws1.cell(row=r, column=2, value=value)
        lc.fill = LBLUE_FILL; vc.fill = fill
        lc.font = bold_font;  vc.font = Font(bold=True, size=11)
        lc.border = vc.border = thin_border
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[r].height = 20

    chart_data_ws = wb.create_sheet("_ChartData", 0)
    chart_data_ws.sheet_state = "hidden"
    chart_data_ws["A1"], chart_data_ws["B1"] = "Risk", "Count"
    for i, lvl in enumerate(["High","Medium","Low"], 2):
        chart_data_ws.cell(row=i, column=1, value=lvl)
        chart_data_ws.cell(row=i, column=2, value=data["risk_counts"].get(lvl,0))
    bar = BarChart()
    bar.type = "col"; bar.title = "Risk Distribution"
    bar.y_axis.title = "Patients"; bar.x_axis.title = "Risk Level"
    bar.width = 14; bar.height = 9
    bar.add_data(Reference(chart_data_ws, min_col=2, min_row=1, max_row=4),
                 titles_from_data=True)
    bar.set_categories(Reference(chart_data_ws, min_col=1, min_row=2, max_row=4))
    ws1.add_chart(bar, "D4")

    ws2 = wb.create_sheet("Risk Predictions")
    make_title(ws2, "Risk Predictions")
    set_header_row(ws2, 4, ["#","Patient ID","Risk Level","Probability (%)","Date"],
                   [5,18,14,18,22])
    for i, p in enumerate(data["predictions"], 1):
        r = 4 + i
        try:
            ws2.cell(row=r, column=1, value=i)
            ws2.cell(row=r, column=2, value=str(p[1]))
            ws2.cell(row=r, column=3, value=str(p[2]))
            ws2.cell(row=r, column=4, value=round(float(p[1])*100,2) if isinstance(p[1],float) else "—")
            ws2.cell(row=r, column=5, value=str(p[3])[:16])
        except Exception:
            for c, v in enumerate(list(p)[:5], 1):
                ws2.cell(row=r, column=c, value=str(v)[:30])
        style_data_row(ws2, r, 5, alt=(i % 2 == 0))

    ws3 = wb.create_sheet("Vitals Monitor")
    make_title(ws3, "Patient Vitals")
    set_header_row(ws3, 4,
                   ["#","Patient ID","Blood Pressure","Heart Rate","Temp (°C)","Blood Sugar","Date"],
                   [5,18,18,14,14,16,22])
    for i, v in enumerate(data["vitals"], 1):
        r = 4 + i
        ws3.cell(row=r, column=1, value=i)
        ws3.cell(row=r, column=2, value=str(v[1]))
        ws3.cell(row=r, column=3, value=str(v[1]))
        ws3.cell(row=r, column=4, value=v[2])
        ws3.cell(row=r, column=5, value=v[3])
        ws3.cell(row=r, column=6, value=v[4])
        ws3.cell(row=r, column=7, value=str(v[5])[:16])
        style_data_row(ws3, r, 7, alt=(i % 2 == 0))
    if len(data["vitals"]) > 1:
        n  = min(len(data["vitals"]), 20)
        lc = LineChart()
        lc.title = "Heart Rate Trend"; lc.y_axis.title = "BPM"
        lc.width = 16; lc.height = 9; lc.smooth = True
        lc.add_data(Reference(ws3, min_col=4, min_row=5, max_row=4+n))
        ws3.add_chart(lc, "I4")

    ws4 = wb.create_sheet("Symptom Alerts")
    make_title(ws4, "Symptom Reports")
    set_header_row(ws4, 4,
                   ["#","Patient ID","Fever","Pain Level","Breathing","Notes","Flagged","Date"],
                   [5,18,10,12,14,30,10,20])
    for i, s in enumerate(data["symptoms"], 1):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i)
        for c, v in enumerate(list(s)[1:8], 2):
            ws4.cell(row=r, column=c, value=str(v) if v is not None else "—")
        style_data_row(ws4, r, 8, alt=(i % 2 == 0))
        flagged_cell = ws4.cell(row=r, column=7)
        if str(flagged_cell.value).lower() in ("1","true","yes"):
            flagged_cell.fill = RED_FILL
            flagged_cell.font = Font(bold=True, color="FFFFFF", size=10)

    ws5 = wb.create_sheet("Patient Feedback")
    make_title(ws5, "Patient Feedback")
    ws5["A3"] = f"Average Rating: {data['avg_rating']} / 5.0"
    ws5["A3"].font = Font(bold=True, size=11, color="34a853")
    set_header_row(ws5, 5, ["#","Patient ID","Rating","Comments","Date"],
                   [5,18,10,50,20])
    for i, f in enumerate(data["feedbacks"], 1):
        r = 5 + i
        ws5.cell(row=r, column=1, value=i)
        ws5.cell(row=r, column=2, value=str(f[1]))
        rating_cell = ws5.cell(row=r, column=3, value=int(f[2]))
        ws5.cell(row=r, column=4, value=str(f[3] or "—"))
        ws5.cell(row=r, column=5, value=str(f[4])[:16] if len(f) > 4 else "—")
        style_data_row(ws5, r, 5, alt=(i % 2 == 0))
        if int(f[2]) >= 4:
            rating_cell.fill = GREEN_FILL
            rating_cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif int(f[2]) <= 2:
            rating_cell.fill = RED_FILL
            rating_cell.font = Font(bold=True, color="FFFFFF", size=10)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action(session['user'], "EXPORT_REPORT_EXCEL")
    response = make_response(buf.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = (
        f'attachment; filename=analytics_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE ④ — EXPORT SINGLE PATIENT PDF
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/export-patient-pdf/<patient_id>")
def export_patient_pdf(patient_id):
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')

    report       = get_patient_report(patient_id)
    medications  = get_patient_medications(patient_id)
    appointments = get_patient_appointments(patient_id)
    vitals       = get_patient_vitals(patient_id)

    buf    = BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.75*inch,  bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    BLUE   = colors.HexColor("#1a73e8")
    LBLUE  = colors.HexColor("#e8f0fe")

    title_s = ParagraphStyle("T",  parent=styles["Title"],
                               textColor=BLUE, fontSize=18, spaceAfter=4)
    h1_s    = ParagraphStyle("H1", parent=styles["Heading1"],
                               textColor=BLUE, fontSize=12, spaceAfter=3)
    small_s = ParagraphStyle("S",  parent=styles["Normal"],
                               fontSize=8, textColor=colors.grey)

    def kv(rows):
        t = Table(rows, colWidths=[2.2*inch, 4*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (0,-1), LBLUE),
            ('FONTNAME',      (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,0), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        return t

    story = []
    pname = report.get("name", patient_id) if report else patient_id
    story.append(Paragraph("Patient Report", title_s))
    story.append(Paragraph(
        f"Patient: {pname}  |  ID: {patient_id}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", small_s))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=12))

    if report:
        story.append(Paragraph("Patient Information", h1_s))
        story.append(kv([
            ["Name",  report.get("name","—")],
            ["Email", report.get("email","—")],
            ["Role",  report.get("role","—")],
        ]))
        story.append(Spacer(1, 10))

    pred = get_latest_prediction(patient_id)
    if pred:
        story.append(Paragraph("Latest Risk Assessment", h1_s))
        story.append(kv([
            ["Risk Level",  pred[0]],
            ["Probability", f"{round(pred[1]*100,2)}%"],
            ["Care Plan",   pred[2][:300] + ("…" if len(pred[2])>300 else "")],
            ["Assessed On", str(pred[3])[:16]],
        ]))
        story.append(Spacer(1, 10))

    if vitals:
        story.append(Paragraph("Recent Vitals", h1_s))
        v_hdr  = [["Date","BP","HR","Temp","B.Sugar"]]
        v_rows = [[str(v[5])[:16],str(v[1]),str(v[2]),str(v[3]),str(v[4])]
                  for v in vitals[:8]]
        vt = Table(v_hdr + v_rows,
                   colWidths=[2*inch,1.2*inch,1*inch,1*inch,1.5*inch])
        vt.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(vt)
        story.append(Spacer(1, 10))

    if medications:
        story.append(Paragraph("Prescribed Medications", h1_s))
        for m in medications:
            story.append(kv([
                ["Medicine",     m[1]],
                ["Dosage",       m[2]],
                ["Frequency",    m[3]],
                ["Duration",     m[4]],
                ["Instructions", m[5] or "As directed"],
            ]))
            story.append(Spacer(1, 6))

    if appointments:
        story.append(Paragraph("Scheduled Appointments", h1_s))
        a_hdr  = [["Date","Time","Reason","Status"]]
        a_rows = [[a[1],a[2],a[3],a[4]] for a in appointments]
        at = Table(a_hdr + a_rows,
                   colWidths=[1.5*inch,1*inch,3*inch,1.2*inch])
        at.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(at)

    doc.build(story)
    buf.seek(0)
    log_action(session['user'], "EXPORT_PATIENT_PDF", f"Patient: {patient_id}")
    resp = make_response(buf.read())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = \
        f'attachment; filename=patient_{patient_id}_report.pdf'
    return resp


# ═══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        uid    = request.form['unique_id']
        pwd    = request.form['password']
        result = login_user(uid, pwd)
        if result:
            role, name = result
            session['user'] = uid
            session['role'] = role
            session['name'] = name
            log_action(uid, "LOGIN", f"Role: {role}")
            return redirect("/dashboard")
        flash("Invalid Login ID or Password", "error")
    return render_template("login.html")


# ═══════════════════════════════════════════════════════════════════════════════
# REGISTER
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name  = request.form['name']
        email = request.form['email']
        pwd   = request.form['password']
        role  = request.form['role']
        uid   = register_user(name, email, pwd, role)
        if uid:
            send_welcome_email(email, name, uid)
            log_action(uid, "REGISTER", f"Role: {role}")
            return render_template("register.html", unique_id=uid)
        flash("Email already exists", "error")
    return render_template("register.html")


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/dashboard")
def dashboard():
    if 'user' not in session:
        return redirect('/')
    if session['role'] in ['doctor', 'hospital']:
        return redirect('/doctor-dashboard')

    uid           = session['user']
    prediction    = get_latest_prediction(uid)
    discharge     = get_discharge_status(uid)
    appointments  = get_patient_appointments(uid)
    medications   = get_patient_medications(uid)
    notifications = get_notifications(uid)
    unread_count  = get_unread_count(uid)
    messages      = get_messages_for_patient(uid)
    unread_msgs   = get_unread_message_count(uid)
    vitals        = get_patient_vitals(uid)

    pred_history    = get_patient_predictions(uid)
    chart_labels    = [p[2] for p in pred_history]
    chart_data      = [round(p[1] * 100, 2) for p in pred_history]
    vitals_reversed = list(reversed(vitals[:10]))
    vital_labels    = [v[5] for v in vitals_reversed]
    vital_hr        = [v[2] for v in vitals_reversed]
    vital_temp      = [v[3] for v in vitals_reversed]
    vital_bs        = [v[4] for v in vitals_reversed]

    return render_template("dashboard.html",
        name=session['name'], prediction=prediction, discharge=discharge,
        appointments=appointments, medications=medications,
        notifications=notifications, unread_count=unread_count,
        messages=messages, unread_msgs=unread_msgs, vitals=vitals,
        chart_labels=chart_labels, chart_data=chart_data,
        vital_labels=vital_labels, vital_hr=vital_hr,
        vital_temp=vital_temp, vital_bs=vital_bs)


@app.route("/mark-notifications-read")
def mark_read():
    if 'user' in session:
        mark_notifications_read(session['user'])
        mark_messages_read(session['user'])
    return redirect('/dashboard')


@app.route("/acknowledge-care-plan", methods=["POST"])
def acknowledge():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    acknowledge_care_plan(session['user'])
    log_action(session['user'], "ACKNOWLEDGE_CARE_PLAN")
    flash("✅ Care plan acknowledged.", "success")
    return redirect('/dashboard')


@app.route("/sos", methods=["POST"])
def sos_alert():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    uid      = session['user']
    name     = session['name']
    vitals   = get_patient_vitals(uid)
    last_v   = vitals[0] if vitals else None
    last_sym = f"BP:{last_v[1]}, HR:{last_v[2]}, Temp:{last_v[3]}" if last_v else "No recent vitals"
    trigger_sos(uid, f"Emergency SOS triggered by {name}")
    send_sos_email(name, uid, last_sym)
    add_notification(uid, "🆘 Your SOS alert has been sent to the medical team.")
    log_action(uid, "SOS_TRIGGERED", f"Patient: {name}")
    flash("🆘 Emergency alert sent! The medical team has been notified.", "error")
    return redirect('/dashboard')


@app.route("/submit-vitals", methods=["POST"])
def submit_vitals():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    uid         = session['user']
    bp          = request.form.get('blood_pressure', '').strip()
    hr          = request.form.get('heart_rate', 0)
    temp        = request.form.get('temperature', 0)
    bs          = request.form.get('blood_sugar', 0)
    is_abnormal = save_vital_signs(uid, bp, hr, temp, bs)
    log_action(uid, "SUBMIT_VITALS", f"BP:{bp}, HR:{hr}, Temp:{temp}, BS:{bs}")
    if is_abnormal:
        patient = get_patient_email(uid)
        if patient:
            send_vital_alert_email(session['name'], bp, hr, temp, bs)
        add_notification(uid, "⚠️ Abnormal vitals detected. Doctor has been alerted.")
        flash("⚠️ Abnormal readings detected. Doctor alerted.", "error")
    else:
        flash("✅ Vitals recorded successfully.", "success")
    return redirect('/dashboard')


@app.route("/submit-symptoms", methods=["POST"])
def submit_symptoms():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    uid        = session['user']
    fever      = request.form.get('fever', 'No')
    pain       = request.form.get('pain_level', 0)
    breathing  = request.form.get('breathing', 'Normal')
    notes      = request.form.get('notes', '').strip()
    is_flagged = save_symptom(uid, fever, pain, breathing, notes)
    log_action(uid, "SUBMIT_SYMPTOMS", f"Fever:{fever}, Pain:{pain}, Breathing:{breathing}")
    if is_flagged:
        send_symptom_alert_email(session['name'], fever, pain, breathing, notes)
        flash("⚠️ Symptoms flagged. Doctor has been alerted.", "error")
    else:
        flash("✅ Symptoms recorded.", "success")
    return redirect('/dashboard')


@app.route("/submit-feedback", methods=["POST"])
def submit_feedback():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    rating   = request.form.get('rating', 5)
    comments = request.form.get('comments', '').strip()
    save_feedback(session['user'], rating, comments)
    log_action(session['user'], "SUBMIT_FEEDBACK", f"Rating: {rating}")
    flash("✅ Thank you for your feedback!", "success")
    return redirect('/dashboard')


@app.route("/download-care-plan")
def download_care_plan():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    uid          = session['user']
    prediction   = get_latest_prediction(uid)
    medications  = get_patient_medications(uid)
    appointments = get_patient_appointments(uid)
    if not prediction:
        flash("No care plan available to download.", "error")
        return redirect('/dashboard')
    lines = ["="*60, "     POST DISCHARGE CARE PLAN REPORT", "="*60,
             f"Patient Name  : {session['name']}", f"Patient ID    : {uid}",
             f"Generated On  : {datetime.now().strftime('%Y-%m-%d %H:%M')}", "",
             "-"*60, "RISK ASSESSMENT", "-"*60,
             f"Risk Level    : {prediction[0]}",
             f"Probability   : {round(prediction[1]*100, 2)}%",
             f"Assessed On   : {prediction[3]}", "",
             "-"*60, "CARE PLAN", "-"*60, prediction[2], ""]
    if medications:
        lines += ["-"*60, "PRESCRIBED MEDICATIONS", "-"*60]
        for m in medications:
            lines += [f"Medicine  : {m[1]}", f"Dosage    : {m[2]}  |  Frequency: {m[3]}",
                      f"Duration  : {m[4]}"]
            if m[5]: lines.append(f"Instructions: {m[5]}")
            lines.append("")
    if appointments:
        lines += ["-"*60, "UPCOMING APPOINTMENTS", "-"*60]
        for a in appointments:
            lines += [f"Date   : {a[1]}  |  Time: {a[2]}",
                      f"Reason : {a[3]}  |  Status: {a[4]}", ""]
    lines += ["="*60, "Post Discharge Care Optimization System", "="*60]
    content  = "\n".join(lines)
    log_action(uid, "DOWNLOAD_CARE_PLAN")
    response = make_response(content)
    response.headers['Content-Type']        = 'text/plain'
    response.headers['Content-Disposition'] = f'attachment; filename=care_plan_{uid}.txt'
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# DOCTOR DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/doctor-dashboard")
def doctor_dashboard():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')

    discharges      = get_pending_discharges()
    flagged_count   = get_flagged_symptom_count()
    analytics       = get_analytics()
    all_patients    = get_all_patients()
    unread_count    = get_unread_count(session['user'])
    sos_count       = get_unresolved_sos_count()
    all_vitals_raw  = get_all_vitals()
    recent_vitals   = sorted(all_vitals_raw[:20],
                              key=lambda v: v[6] if len(v) > 6 else 0,
                              reverse=True)[:10]
    all_sos         = get_all_sos_alerts()
    recent_sos      = all_sos[:5] if all_sos else []
    all_fb          = get_all_feedback()
    recent_feedback = all_fb[:5] if all_fb else []
    avg_rating      = round(sum(f[2] for f in all_fb) / len(all_fb), 1) if all_fb else 0

    a = analytics
    return render_template("doctor_dashboard.html",
        discharges=discharges, flagged_count=flagged_count, analytics=analytics,
        all_patients=all_patients, unread_count=unread_count, sos_count=sos_count,
        risk_labels=['High','Medium','Low'],
        risk_values=[a['high_risk'],a['medium_risk'],a['low_risk']],
        dis_labels=['Approved','Rejected','Pending'],
        dis_values=[a['approved'],a['rejected'],a['pending']],
        recent_vitals=recent_vitals, recent_sos=recent_sos,
        recent_feedback=recent_feedback, avg_rating=avg_rating)


# ═══════════════════════════════════════════════════════════════════════════════
# PATIENT — EXPORT OWN CARE PLAN AS PDF
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/export-my-care-plan-pdf")
def export_my_care_plan_pdf():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    uid          = session['user']
    name         = session['name']
    prediction   = get_latest_prediction(uid)
    medications  = get_patient_medications(uid)
    appointments = get_patient_appointments(uid)
    vitals       = get_patient_vitals(uid)
    if not prediction:
        flash("No care plan available to export.", "error")
        return redirect('/dashboard')

    buf    = BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.75*inch,  bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    BLUE   = colors.HexColor("#1a3c5e")
    LBLUE  = colors.HexColor("#e8f0fe")
    GREEN  = colors.HexColor("#27ae60")
    RED    = colors.HexColor("#e74c3c")
    AMBER  = colors.HexColor("#f39c12")

    title_s = ParagraphStyle("T",  parent=styles["Title"],
                               textColor=BLUE, fontSize=20, spaceAfter=4)
    h1_s    = ParagraphStyle("H1", parent=styles["Heading1"],
                               textColor=BLUE, fontSize=12, spaceAfter=3)
    small_s = ParagraphStyle("S",  parent=styles["Normal"],
                               fontSize=8,  textColor=colors.grey)
    body_s  = ParagraphStyle("B",  parent=styles["Normal"],
                               fontSize=10, leading=16)

    def kv(rows):
        t = Table(rows, colWidths=[2.2*inch, 4.3*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (0,-1), LBLUE),
            ('FONTNAME',      (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,0), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        return t

    story = []
    story.append(Paragraph("Post Discharge Care Plan", title_s))
    story.append(Paragraph(
        f"Patient: {name}  |  ID: {uid}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", small_s))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=16))
    story.append(Paragraph("Patient Details", h1_s))
    story.append(kv([["Patient Name", name], ["Patient ID", uid],
                     ["Report Date", datetime.now().strftime('%Y-%m-%d %H:%M')]]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Risk Assessment", h1_s))
    story.append(kv([["Risk Level",  prediction[0]],
                     ["Probability", f"{round(prediction[1]*100,2)}%"],
                     ["Assessed On", str(prediction[3])[:16]]]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Care Plan Instructions", h1_s))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#dee2e6"), spaceAfter=8))
    for line in prediction[2].split('\n'):
        if line.strip():
            story.append(Paragraph(line.strip(), body_s))
            story.append(Spacer(1, 4))
    story.append(Spacer(1, 10))
    if medications:
        story.append(Paragraph("Prescribed Medications", h1_s))
        for m in medications:
            story.append(kv([["Medicine", m[1]], ["Dosage", m[2]],
                              ["Frequency", m[3]], ["Duration", m[4]],
                              ["Instructions", m[5] or "As directed"]]))
            story.append(Spacer(1, 8))
    if appointments:
        story.append(Paragraph("Scheduled Appointments", h1_s))
        at = Table([["Date","Time","Reason","Status"]] +
                   [[a[1],a[2],a[3],a[4]] for a in appointments],
                   colWidths=[1.4*inch,1*inch,3.2*inch,1*inch])
        at.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(at)
        story.append(Spacer(1, 10))
    if vitals:
        v = vitals[0]
        story.append(Paragraph("Latest Vital Signs", h1_s))
        story.append(kv([["Blood Pressure", str(v[1])],
                          ["Heart Rate",     f"{v[2]} bpm"],
                          ["Temperature",    f"{v[3]} °C"],
                          ["Blood Sugar",    f"{v[4]} mg/dL"],
                          ["Recorded On",    str(v[5])[:16]]]))
        story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        "Post Discharge Care Optimization System — Patient Copy — Confidential",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    log_action(uid, "EXPORT_MY_CARE_PLAN_PDF")
    resp = make_response(buf.read())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = \
        f'attachment; filename=care_plan_{uid}_{datetime.now().strftime("%Y%m%d")}.pdf'
    return resp


@app.route("/send-message", methods=["POST"])
def send_message_route():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    patient_id = request.form.get('patient_id')
    message    = request.form.get('message', '').strip()
    if patient_id and message:
        send_message(session['user'], patient_id, message)
        add_notification(patient_id,
            f"💬 New message from Dr. {session['name']}: {message[:60]}...")
        log_action(session['user'], "SEND_MESSAGE",
                   f"To: {patient_id}, Msg: {message[:50]}")
        flash("✅ Message sent successfully.", "success")
    else:
        flash("Please select a patient and enter a message.", "error")
    return redirect('/doctor-dashboard')


@app.route("/resolve-sos/<int:sos_id>")
def resolve_sos_route(sos_id):
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    resolve_sos(sos_id)
    log_action(session['user'], "RESOLVE_SOS", f"SOS ID: {sos_id}")
    flash("✅ SOS alert marked as resolved.", "success")
    return redirect('/sos-alerts')


@app.route("/sos-alerts")
def sos_alerts_page():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("sos_alerts.html", alerts=get_all_sos_alerts())


@app.route("/vitals-monitor")
def vitals_monitor():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("vitals_monitor.html", vitals=get_all_vitals())


@app.route("/patient-report/<patient_id>")
def patient_report(patient_id):
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    report = get_patient_report(patient_id)
    log_action(session['user'], "VIEW_PATIENT_REPORT", f"Patient: {patient_id}")
    return render_template("patient_report.html", report=report, patient_id=patient_id)


@app.route("/audit-log")
def audit_log():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("audit_log.html", logs=get_audit_log())


@app.route("/view-feedback")
def view_feedback():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    feedbacks = get_all_feedback()
    avg = round(sum(f[2] for f in feedbacks) / len(feedbacks), 1) if feedbacks else 0
    return render_template("feedback.html", feedbacks=feedbacks, avg=avg)


@app.route("/schedule-appointment", methods=["POST"])
def schedule_appointment_route():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    pid    = request.form.get('patient_id')
    date   = request.form.get('appointment_date')
    time   = request.form.get('appointment_time')
    reason = request.form.get('reason', '').strip()
    schedule_appointment(pid, session['user'], date, time, reason)
    add_notification(pid, f"📅 Appointment scheduled on {date} at {time}.")
    patient = get_patient_email(pid)
    if patient:
        send_appointment_email(patient[1], patient[0], date, time, reason)
    log_action(session['user'], "SCHEDULE_APPOINTMENT",
               f"Patient:{pid}, Date:{date}, Time:{time}")
    flash("✅ Appointment scheduled.", "success")
    return redirect('/doctor-dashboard')


@app.route("/prescribe-medication", methods=["POST"])
def prescribe_medication_route():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    pid   = request.form.get('patient_id')
    med   = request.form.get('medicine_name', '').strip()
    dos   = request.form.get('dosage', '').strip()
    freq  = request.form.get('frequency', '').strip()
    dur   = request.form.get('duration', '').strip()
    instr = request.form.get('instructions', '').strip()
    prescribe_medication(pid, session['user'], med, dos, freq, dur, instr)
    add_notification(pid, f"💊 New medication: {med} — {dos}, {freq}.")
    patient = get_patient_email(pid)
    if patient:
        send_medication_email(patient[1], patient[0], med, dos, freq, dur, instr)
    log_action(session['user'], "PRESCRIBE_MEDICATION", f"Patient:{pid}, Med:{med}")
    flash("✅ Medication prescribed.", "success")
    return redirect('/doctor-dashboard')


@app.route("/predict", methods=["GET", "POST"])
def predict():
    if 'user' not in session:
        return redirect('/')
    if session['role'] not in ['doctor', 'hospital']:
        flash("Access restricted", "error")
        return redirect('/dashboard')
    if request.method == "POST":
        pid  = request.form['patient_id']
        age  = int(request.form['age'])
        gen  = int(request.form['gender'])
        tih  = int(request.form['time_in_hospital'])
        inp  = int(request.form['inpatient'])
        eme  = int(request.form['emergency'])
        dia  = int(request.form['diagnoses'])
        X    = np.array([[age, gen, tih, inp, eme, dia]])
        prob = model.predict_proba(X)[0][1]
        risk = ("High" if prob >= 0.7 else "Medium" if prob >= 0.4 else "Low")
        plan = get_care_plan(risk)
        save_prediction(pid, risk, prob, plan)
        add_notification(pid, f"🧠 Risk assessment: {risk} ({round(prob*100,2)}%)")
        patient = get_patient_email(pid)
        if patient:
            send_risk_alert_email(patient[1], patient[0], risk, prob, plan)
        log_action(session['user'], "PREDICT",
                   f"Patient:{pid}, Risk:{risk}, Prob:{round(prob*100,2)}%")
        return render_template("result.html", risk=risk, prob=round(prob*100,2), plan=plan)
    return render_template("predict.html")


@app.route("/prediction-history")
def prediction_history():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("prediction_history.html", predictions=get_all_predictions())


@app.route("/discharge-history")
def discharge_history():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("discharge_history.html", history=get_all_discharge_history())


@app.route("/symptom-alerts")
def symptom_alerts():
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    return render_template("symptom_alerts.html", symptoms=get_all_symptoms())


@app.route("/my-history")
def my_history():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    preds = [p for p in get_all_predictions() if p[1] == session['user']]
    return render_template("patient_history.html", predictions=preds, name=session['name'])


# ═══════════════════════════════════════════════════════════════════════════════
# FORGOT PASSWORD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get('email', '').strip()
        user  = get_user_by_email(email)
        if user:
            uid, name = user
            token  = secrets.token_urlsafe(32)
            expiry = datetime.now() + timedelta(minutes=30)
            reset_tokens[token] = (uid, expiry)
            link = request.host_url.rstrip('/') + f"/reset-password/{token}"
            send_reset_email(email, name, link)
            flash("✅ Password reset link sent to your email!", "success")
        else:
            flash("✅ If this email is registered, a reset link was sent.", "success")
    return render_template("forgot_password.html")


# ═══════════════════════════════════════════════════════════════════════════════
# RESET PASSWORD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    data = reset_tokens.get(token)
    if not data:
        flash("Invalid link.", "error")
        return redirect("/forgot-password")
    uid, expiry = data
    if datetime.now() > expiry:
        reset_tokens.pop(token, None)
        flash("Link expired.", "error")
        return redirect("/forgot-password")
    if request.method == "POST":
        pwd = request.form.get('password', '')
        cfm = request.form.get('confirm_password', '')
        if pwd != cfm:
            flash("Passwords do not match.", "error")
            return render_template("reset_password.html", token=token)
        update_user_password(uid, pwd)
        reset_tokens.pop(token, None)
        flash("✅ Password updated successfully!", "success")
        return redirect("/")
    return render_template("reset_password.html", token=token)


# ═══════════════════════════════════════════════════════════════════════════════
# PROFILE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/profile", methods=["GET", "POST"])
def profile():
    if 'user' not in session:
        return redirect('/')
    user_profile = get_user_profile(session['user'])
    if request.method == "POST":
        name  = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        pwd   = request.form.get('password', '').strip()
        cfm   = request.form.get('confirm_password', '').strip()
        if pwd and pwd != cfm:
            flash("Passwords do not match.", "error")
            return render_template("profile.html", user=user_profile)
        ok = update_user_profile(
            unique_id=session['user'],
            name=name  or user_profile['name'],
            email=email or user_profile['email'],
            password=pwd if pwd else None
        )
        if ok:
            session['name'] = name or user_profile['name']
            flash("✅ Profile updated.", "success")
            return redirect('/profile')
        flash("Email already in use.", "error")
    return render_template("profile.html", user=user_profile)


@app.route("/request-discharge")
def patient_request_discharge():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    discharge = get_discharge_status(session['user'])
    if discharge and discharge[4] == 0:
        flash("⚠️ Acknowledge your care plan first.", "error")
        return redirect('/dashboard')
    request_discharge(session['user'])
    log_action(session['user'], "REQUEST_DISCHARGE")
    flash("Discharge request submitted.", "success")
    return redirect('/dashboard')


@app.route("/discharge-action/<int:rid>/<status>", methods=["POST"])
def discharge_action(rid, status):
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return redirect('/dashboard')
    remarks = request.form.get('remarks', '').strip()
    update_discharge(rid, status.upper(), session['user'], remarks)
    conn = sqlite3.connect("database.db")
    row  = conn.execute("""
        SELECT u.unique_id, u.name, u.email
        FROM discharge_requests dr
        JOIN users u ON dr.patient_id=u.unique_id
        WHERE dr.id=?
    """, (rid,)).fetchone()
    conn.close()
    if row:
        pid, pname, pemail = row
        send_discharge_decision_email(pemail, pname, status.upper(), remarks)
        add_notification(pid,
            f"{'✅' if status.upper()=='APPROVED' else '❌'} "
            f"Discharge {status.upper()}."
            f"{' Remarks: '+remarks if remarks else ''}")
        log_action(session['user'], f"DISCHARGE_{status.upper()}",
                   f"Patient:{pid}, Remarks:{remarks}")
    flash(f"Discharge {status.capitalize()} successfully.", "success")
    return redirect('/doctor-dashboard')


# ═══════════════════════════════════════════════════════════════════════════════
# HEALTH CHATBOT (GEMINI)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/chatbot")
def chatbot():
    if 'user' not in session or session['role'] != 'patient':
        return redirect('/dashboard')
    return render_template("chatbot.html", name=session['name'])


@app.route("/chatbot-api", methods=["POST"])
def chatbot_api():
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    user_message = request.json.get("message", "").strip()
    if not user_message:
        return jsonify({"error": "Empty message"}), 400
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            system_instruction="""You are a helpful post-discharge health assistant.
            Help patients understand their symptoms, medications, and recovery process.
            Always remind patients to consult their doctor for serious concerns.
            Keep responses concise, friendly and easy to understand."""
        )
        response = gemini_model.generate_content(user_message)
        reply    = response.text
        conn = sqlite3.connect("database.db")
        conn.execute("""
            INSERT INTO chat_history (patient_id, message, response, timestamp)
            VALUES (?, ?, ?, ?)
        """, (session['user'], user_message, reply,
              datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        return jsonify({"response": reply})
    except Exception as e:
        print(f"[CHATBOT ERROR] {e}")
        return jsonify({"response": "Sorry, I'm unable to respond right now. Please try again later."})


# ═══════════════════════════════════════════════════════════════════════════════
# ★ NEW — PREDICTIVE PATIENT DETERIORATION HEATMAP
# ═══════════════════════════════════════════════════════════════════════════════

def _compute_deterioration_score(risk, prob, trend, days_inactive,
                                  has_fever, high_pain, abnormal_breathing,
                                  vitals_abnormal):
    """
    Compute a 0–100 composite deterioration score for sorting / badge colour.

    Weights:
      - Risk probability      : up to 40 pts
      - Trend direction       : ↑ +15, → +5, ↓ +0
      - Inactivity            : >5 days +20, 3-5 days +10, 1-2 days +5
      - Fever                 : +10
      - High pain (>=7)       : +8
      - Abnormal breathing    : +7
      - Abnormal vitals       : +10  (capped at 100 total)
    """
    score = 0
    score += min(prob * 40, 40)                          # probability component
    score += {"↑": 15, "→": 5, "↓": 0}.get(trend, 5)   # trend
    if days_inactive is not None:
        if days_inactive > 5:   score += 20
        elif days_inactive > 2: score += 10
        elif days_inactive > 0: score += 5
    else:
        score += 20                                       # never submitted = worst case
    if has_fever:            score += 10
    if high_pain:            score += 8
    if abnormal_breathing:   score += 7
    if vitals_abnormal:      score += 10
    return min(round(score, 1), 100)


@app.route("/deterioration-heatmap")
def deterioration_heatmap():
    """
    Predictive Patient Deterioration Heatmap.
    Combines risk prediction, vitals, symptoms and inactivity into one
    colour-coded card grid so doctors get a single-screen morning briefing.
    """
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        flash("Unauthorized", "error")
        return redirect('/dashboard')

    patients     = get_all_patients()
    all_symptoms = get_all_symptoms()
    patient_data = []

    # Build a quick symptom lookup: patient_id → latest symptom row
    symptom_map = {}
    for s in all_symptoms:
        pid_sym = str(s[1])
        if pid_sym not in symptom_map:
            symptom_map[pid_sym] = s

    for p in patients:
        uid  = str(p[0])
        name = p[1] if len(p) > 1 else uid

        # ── latest prediction ──────────────────────────────────────────────
        pred         = get_latest_prediction(uid)
        risk         = pred[0] if pred else "Unknown"
        prob         = pred[1] if pred else 0.0          # 0.0–1.0
        prob_pct     = round(prob * 100, 1)

        # ── prediction trend (compare latest two predictions) ──────────────
        pred_history = get_patient_predictions(uid)
        trend        = "→"
        if len(pred_history) >= 2:
            diff = float(pred_history[0][1]) - float(pred_history[1][1])
            trend = "↑" if diff > 0.05 else ("↓" if diff < -0.05 else "→")

        # ── vitals ─────────────────────────────────────────────────────────
        vitals          = get_patient_vitals(uid)
        days_inactive   = None
        vitals_abnormal = False
        last_vitals_str = "Never"

        if vitals:
            latest_v = vitals[0]
            try:
                last_dt         = datetime.strptime(str(latest_v[5])[:16], "%Y-%m-%d %H:%M")
                days_inactive   = (datetime.now() - last_dt).days
                last_vitals_str = last_dt.strftime("%d %b %Y, %H:%M")
            except Exception:
                pass
            # flag abnormal vitals (HR >100 or <50, Temp >38 or <35.5, BS >200)
            try:
                hr   = float(latest_v[2])
                temp = float(latest_v[3])
                bs   = float(latest_v[4])
                if hr > 100 or hr < 50 or temp > 38.0 or temp < 35.5 or bs > 200:
                    vitals_abnormal = True
            except (ValueError, TypeError):
                pass

        # ── latest symptoms ────────────────────────────────────────────────
        sym               = symptom_map.get(uid)
        has_fever         = False
        high_pain         = False
        abnormal_breathing= False

        if sym:
            has_fever          = str(sym[2]).lower() == "yes" if len(sym) > 2 else False
            try:
                high_pain      = int(sym[3]) >= 7 if len(sym) > 3 else False
            except (ValueError, TypeError):
                high_pain      = False
            breathing_val      = str(sym[4]).strip() if len(sym) > 4 else "Normal"
            abnormal_breathing = breathing_val not in ["Normal", "None", "", "No"]

        # ── composite deterioration score ──────────────────────────────────
        det_score = _compute_deterioration_score(
            risk, prob, trend, days_inactive,
            has_fever, high_pain, abnormal_breathing, vitals_abnormal
        )

        # ── alert badge: what is triggering concern ────────────────────────
        alerts = []
        if risk == "High":                                   alerts.append("🔴 High Risk")
        if trend == "↑":                                     alerts.append("📈 Worsening Trend")
        if days_inactive is None or days_inactive > 5:       alerts.append("💤 Inactive 5+ Days")
        elif days_inactive > 2:                              alerts.append("💤 Inactive 3+ Days")
        if has_fever:                                        alerts.append("🌡️ Fever")
        if high_pain:                                        alerts.append("😣 High Pain")
        if abnormal_breathing:                               alerts.append("🫁 Abnormal Breathing")
        if vitals_abnormal:                                  alerts.append("⚠️ Abnormal Vitals")

        patient_data.append({
            "uid":               uid,
            "name":              name,
            "risk":              risk,
            "prob":              prob_pct,
            "trend":             trend,
            "days_inactive":     days_inactive,
            "last_vitals":       last_vitals_str,
            "has_fever":         has_fever,
            "high_pain":         high_pain,
            "abnormal_breathing":abnormal_breathing,
            "vitals_abnormal":   vitals_abnormal,
            "det_score":         det_score,
            "alerts":            alerts,
        })

    # Sort: highest deterioration score first
    patient_data.sort(key=lambda x: -x["det_score"])

    # Summary counts for the top stats bar
    critical_count = sum(1 for p in patient_data if p["det_score"] >= 70)
    warning_count  = sum(1 for p in patient_data if 40 <= p["det_score"] < 70)
    stable_count   = sum(1 for p in patient_data if p["det_score"] < 40)
    inactive_count = sum(1 for p in patient_data
                         if p["days_inactive"] is None or p["days_inactive"] > 3)

    log_action(session['user'], "VIEW_DETERIORATION_HEATMAP")
    return render_template(
        "deterioration_heatmap.html",
        patients        = patient_data,
        critical_count  = critical_count,
        warning_count   = warning_count,
        stable_count    = stable_count,
        inactive_count  = inactive_count,
        generated_at    = datetime.now().strftime("%d %b %Y, %H:%M"),
    )


@app.route("/heatmap-api")
def heatmap_api():
    """
    JSON endpoint — called by the heatmap page every 60 s for live refresh
    without a full page reload.
    """
    if 'user' not in session or session['role'] not in ['doctor', 'hospital']:
        return jsonify({"error": "Unauthorized"}), 401

    patients     = get_all_patients()
    all_symptoms = get_all_symptoms()
    symptom_map  = {}
    for s in all_symptoms:
        pid_sym = str(s[1])
        if pid_sym not in symptom_map:
            symptom_map[pid_sym] = s

    result = []
    for p in patients:
        uid          = str(p[0])
        pred         = get_latest_prediction(uid)
        risk         = pred[0] if pred else "Unknown"
        prob         = pred[1] if pred else 0.0
        pred_history = get_patient_predictions(uid)
        trend        = "→"
        if len(pred_history) >= 2:
            diff  = float(pred_history[0][1]) - float(pred_history[1][1])
            trend = "↑" if diff > 0.05 else ("↓" if diff < -0.05 else "→")

        vitals          = get_patient_vitals(uid)
        days_inactive   = None
        vitals_abnormal = False
        if vitals:
            try:
                last_dt       = datetime.strptime(str(vitals[0][5])[:16], "%Y-%m-%d %H:%M")
                days_inactive = (datetime.now() - last_dt).days
            except Exception:
                pass
            try:
                hr   = float(vitals[0][2])
                temp = float(vitals[0][3])
                bs   = float(vitals[0][4])
                if hr > 100 or hr < 50 or temp > 38.0 or temp < 35.5 or bs > 200:
                    vitals_abnormal = True
            except (ValueError, TypeError):
                pass

        sym                = symptom_map.get(uid)
        has_fever          = str(sym[2]).lower() == "yes" if sym and len(sym) > 2 else False
        try:
            high_pain      = int(sym[3]) >= 7 if sym and len(sym) > 3 else False
        except (ValueError, TypeError):
            high_pain      = False
        breathing_val      = str(sym[4]).strip() if sym and len(sym) > 4 else "Normal"
        abnormal_breathing = breathing_val not in ["Normal", "None", "", "No"]

        det_score = _compute_deterioration_score(
            risk, prob, trend, days_inactive,
            has_fever, high_pain, abnormal_breathing, vitals_abnormal
        )
        result.append({
            "uid":       uid,
            "name":      p[1] if len(p) > 1 else uid,
            "risk":      risk,
            "prob":      round(prob * 100, 1),
            "trend":     trend,
            "det_score": det_score,
        })

    result.sort(key=lambda x: -x["det_score"])
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════════════════
# LOGOUT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/logout")
def logout():
    if 'user' in session:
        log_action(session['user'], "LOGOUT")
    session.clear()
    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)